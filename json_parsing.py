import os
import json
import openpyxl
from operator import itemgetter


def parse_headers(json_data, ws):
    headers = []
    for header in json_data['headers']:
        info = header['properties']
        usefull_info = {
            'width': int(info['Width']),
            'x': int(info['X']),
            'info': info['QuickInfo']
        }
        headers.append(usefull_info)
    # Сортируем по значению x, чтобы соблюсти правильный порядок колонок
    headers.sort(key=itemgetter('x'))
    # Поле width оказалось неактуальным для ширины колонок в экселе, так что ширина колонок остается стандартной
    ws.append(header['info'] for header in headers)
    # Возвращаем пронумерованные координаты X для заголовков (колонка в экселе - координата) для заполнения значений
    enum_headers = enumerate((header['x'] for header in headers), start=1)
    ret_headers = dict((j, i) for i, j in enum_headers)

    return ret_headers


def parse_values(json_data, ws, headers_to_cols):
    """
    Функция заполнения ячеек. Имеет слудующие особенности:
        - проверка, что X ячейки соотвествует X одного из заголовков (1 заголовок - одно значение при неизменном Y);
        - если в исходном файле количество значений не делится нацело на количество заголовков, программа оставляет
        отсутствующие значения пустыми, сообщений об ошибках или предупреждений на такой случай не предусмотрено
    """
    values = []
    for value in json_data['values']:
        info = value['properties']
        usefull_info = {
            'x': int(info['X']),
            'y': int(info['Y']),
            'info': info['Text']
        }
        values.append(usefull_info)
    values.sort(key=itemgetter('y', 'x'))
    y0 = values[0]['y']
    curr_row = 2
    # Создаём копию переданного словаря колонок, чтобы можно было удалять использованные для текущей строки значения
    curr_cols = headers_to_cols.copy()
    for value in values:
        if value['y'] > y0:
            curr_row += 1
            y0 = value['y']
            # Обновляем текущий словарь колонок при переходе на следующую строку
            curr_cols = headers_to_cols.copy()
        if value['x'] in curr_cols:
            # Используем оператор pop, чтобы исключить наложение ячеек друг на друга
            ws.cell(curr_row, curr_cols.pop(value['x']), value['info'])
        else:
            raise Exception(f'Данное значение X value ({value["x"]}) не подходит по полю X ни одному Header. Проверьте '
                            f'валидность данных')


if __name__ == '__main__':
    # Задаём путь, где брать json
    json_location = 'json_test_files'
    # Создаем новую книгу
    wb = openpyxl.Workbook()
    # Удаляем созданную автоматически страницу, мы будем генерировать их в цикле для каждого файла
    wb.remove(wb.active)

    for json_file in os.listdir(json_location):
        if json_file.endswith('.json'):
            curr_ws = wb.create_sheet()
            with open(os.path.join(json_location, json_file), 'r', encoding='utf-8') as curr_file:
                raw_data = json.load(curr_file)
                curr_headers = parse_headers(raw_data, curr_ws)
                parse_values(raw_data, curr_ws, curr_headers)

    wb.save('FileReport.xlsx')
    print('Excel файл сгенерирован')
